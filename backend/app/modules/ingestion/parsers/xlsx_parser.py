"""XLSX 解析（openpyxl）：每个 sheet 一页；首行表头；合并单元格向下填充。"""

from __future__ import annotations

import io

from app.modules.ingestion.parsers.common import page_from_blocks, text_block
from app.modules.ingestion.parsers.pdf import PageParse


def parse_xlsx_bytes(data: bytes) -> list[PageParse]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    pages: list[PageParse] = []
    try:
        for sheet_idx, ws in enumerate(wb.worksheets, start=1):
            grid = _sheet_grid(ws)
            if not grid:
                pages.append(page_from_blocks(sheet_idx, []))
                continue
            title = f"工作表：{ws.title}"
            md = _grid_to_markdown(grid)
            blocks = [
                text_block(title, order=0, level=1, size=14.0, bold=True),
                text_block(md, order=1, size=11.0, block_type="table"),
            ]
            pages.append(page_from_blocks(sheet_idx, blocks))
    finally:
        wb.close()
    return pages or [page_from_blocks(1, [])]


def _sheet_grid(ws: object) -> list[list[str]]:
    from openpyxl.worksheet.worksheet import Worksheet

    assert isinstance(ws, Worksheet)
    # 合并单元格：先建值表再向下/向右填充左上角值
    values: dict[tuple[int, int], str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            values[(cell.row, cell.column)] = str(cell.value).strip()

    for merged in ws.merged_cells.ranges:
        top_left = values.get((merged.min_row, merged.min_col), "")
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                if (r, c) not in values and top_left:
                    values[(r, c)] = top_left

    if not values:
        return []
    max_r = max(r for r, _ in values)
    max_c = max(c for _, c in values)
    grid: list[list[str]] = []
    for r in range(1, max_r + 1):
        grid.append([values.get((r, c), "") for c in range(1, max_c + 1)])
    # 去掉全空行
    return [row for row in grid if any(cell.strip() for cell in row)]


def _grid_to_markdown(grid: list[list[str]]) -> str:
    if not grid:
        return ""
    width = max(len(r) for r in grid)
    norm = [r + [""] * (width - len(r)) for r in grid]
    header = norm[0]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for r in norm[1:]:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)
